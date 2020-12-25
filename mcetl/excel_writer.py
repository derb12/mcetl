# -*- coding: utf-8 -*-
"""ExcelWriterHandler class, used to safely open and close files and apply styles.

@author  : Donald Erb
Created on Dec 9, 2020

Notes
-----
openpyxl is imported within methods of ExcelWriterHandler in order
to reduce the import time of mcetl since the writer is only needed
if writing to Excel.

"""


from pathlib import Path
import traceback

import pandas as pd
import PySimpleGUI as sg

from .utils import PROCEED_COLOR
#openpyxl is imported within methods of ExcelWriterHandler


class ExcelWriterHandler:
    """
    A helper for pandas's ExcelWriter for opening/saving files and applying styles.

    Parameters
    ----------
    file_name : str or Path
        The file name or path for the Excel file to be created.
    new_file : bool, optional
        If False (default), will append to an existing file. If True, or if
        no file currently exists with the given file_name, a new file will
        be created, even if a file with the same name currently exists.
    styles : dict(dict or openpyxl.styles.NamedStyle)
        A dictionary of either nested dictionaries, used to create openpyxl
        NamedStyle objects, or NamedStyle objects. All styles in the
        dictionary will be added to the Excel workbook if they do not
        currently exist in the workbook.
    writer : pd.ExcelWriter or None
        The ExcelWriter (_OpenpyxlWriter from pandas) used for writing
        to Excel. If it is a pd.ExcelWriter, its engine must be "openpyxl".
    **kwargs
        Any additional keyword arguments to pass to pd.ExcelWriter.

    Attributes
    ----------
    loaded : bool
        If True, designates that the class is done initializing, which
        if appending to an existing Excel file means that the file is
        ready to write to. Useful for initializing within a separate thread.
    styles : dict(dict)
        A nested dictionary of dictionaries, used to create openpyxl
        NamedStyle objects to include in self.writer.book. The styles
        are used as a class attribute to ensure that the necessary
        NamedStyles are always included in the Excel book within mcetl.
    writer : pd.ExcelWriter
        The ExcelWriter (_OpenpyxlWriter from pandas) used for writing
        to Excel.

    Notes
    -----
    Either file_name or writer must be specified at initialization.

    This class is used for ensuring that an existing file is closed before
    saving and/or writing, if appending, and that the pd.ExcelWriter
    has all of the necessary NamedStyle objects. Otherwise, most calls should
    be to the writer attribute.

    """

    styles = {
        'fitting_header_even': {
            'font': dict(size=12, bold=True),
            'fill': dict(fill_type='solid', start_color='F9B381', end_color='F9B381'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_header_odd': {
            'font': dict(size=12, bold=True),
            'fill': dict(fill_type='solid', start_color='73A2DB', end_color='73A2DB'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_subheader_even': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_subheader_odd': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'),
            'border': dict(bottom=dict(style='thin')),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True)
        },
        'fitting_columns_even': {
            'fill': dict(fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'),
            'alignment': dict(horizontal='center', vertical='center'),
            'number_format': '0.00'
        },
        'fitting_columns_odd': {
            'fill': dict(fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'),
            'alignment': dict(horizontal='center', vertical='center'),
            'number_format': '0.00'
        },
        'fitting_descriptors_even': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='FFEAD6', end_color='FFEAD6'),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True),
            'number_format': '0.000'
        },
        'fitting_descriptors_odd': {
            'font': dict(bold=True),
            'fill': dict(fill_type='solid', start_color='DBEDFF', end_color='DBEDFF'),
            'alignment': dict(horizontal='center', vertical='center', wrap_text=True),
            'number_format': '0.000'
        }
    }

    def __init__(self, file_name=None, new_file=False, styles=None, writer=None, **kwargs):
        """
        Raises
        ------
        TypeError
            Raised if both file_name and writer are None.
        ValueError
            Raised if the engine of the input writer is not "openpyxl".

        """

        self.loaded = False # used for multithreading
        if file_name is None and writer is None:
            raise TypeError(
                'Both file_name and writer cannot be None when creating an ExcelWriterHandler.'
            )
        elif writer is None:
            self.writer = self._create_writer(file_name, new_file, **kwargs)
        elif writer.engine != 'openpyxl':
            raise ValueError(
                'ExcelWriterHandler can only use ExcelWriters with "openpyxl" engine.'
            )
        else:
            self.writer = writer

        self.add_styles(styles)
        self.loaded = True


    def __str__(self):
        return f'{self.__class__.__name__}(path={self.writer.path})'


    def _create_writer(self, file_name, new_file, **kwargs):
        """
        Creates the pandas ExcelWriter.

        Parameters
        ----------
        file_name : str or Path
            The file name or path for the Excel file to be created.
        new_file : bool
            If False (default), will append to an existing file. If True, or if
            no file currently exists with the given file_name, a new file will
            be created, even if a file with the same name currently exists.
        **kwargs
            Any additional keyword arguments to pass to pd.ExcelWriter.

        Notes
        -----
        If appending to a file, makes the user close the file before proceeding
        because any further changes to the file after creating the ExcelWriter
        will be lost.

        """

        path = Path(file_name)
        if new_file or not path.exists():
            mode = 'w'
        else:
            mode = 'a'
            while True:
                try:
                    path.rename(path) # errors if file is currently open
                except PermissionError:
                    sg.popup_ok(
                        (f'Please close {path.name} so it can be opened in'
                         'python.\nUntil the file is saved in python, any '
                         'additional\nchanges made by the user will be lost.\n'),
                        title='Close File'
                    )
                else:
                    break

        return pd.ExcelWriter(file_name, engine='openpyxl', mode=mode, **kwargs)


    def add_styles(self, input_styles=None):
        """
        Adds NamedStyles to the Excel workbook.

        Ensures that at least self.styles are added to the Excel workbook.

        Parameters
        ----------
        input_styles : dict(dict or openpyxl.styles.NamedStyle), optional
            A dictionary of either nested dictionaries, used to create openpyxl
            NamedStyle objects, or NamedStyle objects. All styles in the
            dictionary will be added to the Excel workbook if they do not
            currently exist in the workbook. The keys of the dictionary will
            be the named of the created NamedStyle within the Excel workbook.

        """

        styles = input_styles if input_styles is not None else {}
        for key, value in self.styles.items():
            if key not in styles:
                styles[key] = value

        for name, style in styles.items():
            if name not in self.writer.book.named_styles:
                named_style = self._create_named_style(name, style)
                self.writer.book.add_named_style(named_style)


    @classmethod
    def _create_named_style(cls, name, style):
        """
        Creates an openpyxl NamedStyle from the input style.

        Parameters
        ----------
        name : str
            The desired name of the NamedStyle.
        style : dict or openpyxl.styles.NamedStyle
            If the input is already a NamedStyle, it is returned. Otherwise,
            will convert a dictionary of values into the necessary keyword
            arguments to create a NamedStyle.

        Returns
        -------
        openpyxl.styles.NamedStyle
            The created NamedStyle object.

        Raises
        ------
        ValueError
            Raised if the input style is a NamedStyle, but its
            name is not equal to the input name. Not strictly necessary,
            but meant to ensure that self.writer has at least the named_styles
            defined in self.styles.

        Notes
        -----
        Could use the _convert_to_style_kwargs method for pandas's _OpenpyxlWriter,
        but it is a private method, so don't want to rely on it.

        """

        from openpyxl.styles import Alignment, NamedStyle, Protection

        if isinstance(style, NamedStyle):
            if name != style.name:
                raise ValueError(
                    'NamedStyle objects must have the same name as its dictionary key.'
                )
            else:
                return style

        borders = cls._openpyxl_border(style.get('border', {}))
        font = cls._openpyxl_font(style.get('font', {}))
        fill = cls._openpyxl_fill(style.get('fill', {}))

        kwargs = {
            'name': name,
            'alignment': Alignment(**style.get('alignment', {})),
            'border': borders,
            'fill': fill,
            'font': font,
            'number_format': style.get('number_format', None),
            'protection': Protection(**style.get('protection', {})),
        }
        other_kwargs = {k: v for k, v in style.items() if k not in kwargs}

        return NamedStyle(**kwargs, **other_kwargs)


    @classmethod
    def _openpyxl_color(cls, values):
        from openpyxl.styles import Color

        if isinstance(values, str):
            return Color(values)
        else:
            return Color(**values)


    @classmethod
    def _openpyxl_font(cls, values):
        from openpyxl.styles import Font

        if isinstance(values, str):
            return Font(values)

        kwargs = {}
        for key, value in values.items():
            if key == 'color':
                kwargs[key] = cls._openpyxl_color(value)
            else:
                kwargs[key] = value

        return Font(**kwargs)


    @classmethod
    def _openpyxl_side(cls, values):
        from openpyxl.styles import Side

        if isinstance(values, str):
            return Side(values)

        kwargs = {}
        for key, value in values.items():
            if key == 'color':
                kwargs[key] = cls._openpyxl_color(value)
            else:
                kwargs[key] = value

        return Side(**kwargs)


    @classmethod
    def _openpyxl_border(cls, values):
        from openpyxl.styles import Border

        kwargs = {}
        for key, value in values.items():
            if key in ('left', 'right', 'top', 'bottom', 'diagonal'):
                kwargs[key] = cls._openpyxl_side(value)
            else:
                kwargs[key] = value

        return Border(**kwargs)


    @classmethod
    def _openpyxl_fill(cls, values):
        from openpyxl.styles import GradientFill, PatternFill

        gradient_kwargs = {}
        pattern_kwargs = {}
        for key, value in values.items():
            gradient_key = pattern_key = val = None
            if key in ('fgColor', 'bgColor', 'start_color', 'end_color'):
                pattern_key = key
                val = cls._openpyxl_color(value)
            if key in ('patternType, fill_type'):
                pattern_key = key
            if key in ('type', 'fill_type'):
                gradient_key = 'type' # GradientFill does not take fill_type key
            if key == 'stop':
                gradient_key = key
                val = [cls._openpyxl_color(v) for v in value]
            if val is None:
                val = value

            if pattern_key:
                pattern_kwargs[pattern_key] = val
            if gradient_key:
                gradient_kwargs[gradient_key] = val
            if pattern_key is None and gradient_key is None:
                # all PatternFill keys should be covered above, but assign
                # anyway to cover unforseen differences in openpyxl versions
                pattern_kwargs[key] = val
                gradient_kwargs[key] = val

        try:
            output = PatternFill(**pattern_kwargs)
        except (TypeError, ValueError):
            output = GradientFill(**gradient_kwargs)

        return output


    def save_excel_file(self):
        """Handles saving the Excel file and the various exceptions that can occur."""

        path = Path(self.writer.path)
        # Ensures that the folder destination exists
        path.parent.mkdir(parents=True, exist_ok=True)
        while True:
            try:
                self.writer.save()
                print('\nSaved Excel file.')
                break

            except PermissionError:
                window = sg.Window(
                    'Save Error',
                    layout=[
                        [sg.Text((f'Trying to overwrite {path.name}.\n\n'
                                'Please close the file and press Proceed'
                                ' to save.\nPress Discard to not save.\n'))],
                        [sg.Button('Discard'),
                        sg.Button('Proceed', button_color=PROCEED_COLOR)]
                    ]
                )
                response = window.read()[0]
                window.close()
                window = None
                if response == 'Discard':
                    break


    @classmethod
    def test_excel_styles(cls, styles):
        """
        Tests whether the input styles create valid Excel styles with openpyxl.

        Parameters
        ----------
        styles : dict(str : dict or openpyxl.styles.NamedStyle)
            The dictionary of styles to test. Values in the dictionary can
            either be nested dictionaries with the necessary keys and values
            to create an openpyxl NamedStyle, or openpyxl.styles.NamedStyle
            objects.

        Returns
        -------
        success : bool
            Returns True if all input styles successfully create openpyxl
            NamedStyle objects; otherwise, returns False.

        """

        failed_styles = []
        for key, style in styles.items():
            try:
                cls._create_named_style(key, style)
            except:
                failed_styles.append((key, traceback.format_exc()))

        if not failed_styles:
            success = True
            print('All input styles were successful.')
        else:
            success = False
            print('The following input styles were incorrect:\n')
            for failure, traceback_message in failed_styles:
                print(f'\nKey: {failure}\n{traceback_message}')

        return success
